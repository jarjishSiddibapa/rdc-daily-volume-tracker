"""Excel import/export service for targets and reports."""

import io
import logging
from datetime import date

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app import db
from app.models import PlantMonthlyTarget, Plant

logger = logging.getLogger(__name__)


def parse_target_excel(file_stream, month_date: date) -> dict:
    """
    Parse an uploaded Excel file with monthly targets.

    Expected columns: plant_code, target_volume
    (Column names are case-insensitive and flexible.)

    Returns dict with status, records saved, and any errors.
    """
    try:
        df = pd.read_excel(file_stream, engine="openpyxl")
    except Exception as exc:
        return {"status": "error", "message": f"Could not read Excel file: {str(exc)}"}

    # Normalize column names
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # Find plant_code column
    code_col = None
    for candidate in ["plant_code", "plantcode", "code", "plant"]:
        if candidate in df.columns:
            code_col = candidate
            break

    # Find target column
    target_col = None
    for candidate in ["target_volume", "target", "targetvolume", "volume", "budget"]:
        if candidate in df.columns:
            target_col = candidate
            break

    if code_col is None:
        return {"status": "error", "message": "Could not find 'plant_code' column in Excel"}
    if target_col is None:
        return {"status": "error", "message": "Could not find 'target_volume' column in Excel"}

    saved = 0
    errors = []

    for _, row in df.iterrows():
        plant_code = str(row[code_col]).strip()

        # Skip blank rows and merged region-header rows (non-numeric target column)
        if not plant_code or plant_code.lower() in ("nan", "none", ""):
            continue

        try:
            target_value = float(row[target_col])
        except (ValueError, TypeError):
            # Region header rows have no numeric target — silently skip them
            continue

        if target_value < 0:
            errors.append(f"{plant_code}: target volume cannot be negative")
            continue

        # Check plant exists
        plant = db.session.get(Plant, plant_code)
        if not plant:
            errors.append(f"{plant_code}: plant not found")
            continue

        # Upsert target
        existing = PlantMonthlyTarget.query.filter_by(
            plant_code=plant_code, month_date=month_date
        ).first()

        if existing:
            existing.target_volume = target_value
        else:
            target = PlantMonthlyTarget(
                plant_code=plant_code,
                month_date=month_date,
                target_volume=target_value,
            )
            db.session.add(target)
        saved += 1

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return {"status": "error", "message": f"Database error: {str(exc)}"}

    return {
        "status": "success",
        "saved": saved,
        "errors": errors,
        "month": month_date.isoformat(),
    }


def export_report_excel(report_data: dict) -> io.BytesIO:
    """
    Export the report data to a formatted Excel file.

    Returns a BytesIO buffer with the .xlsx content.
    """
    output = io.BytesIO()
    wb = None

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Volume Report"

        meta = report_data["meta"]

        # ── Styles ─────────────────────────────────────────────────────────
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        region_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        region_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

        total_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(horizontal="right")
        name_align = Alignment(horizontal="left")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ── Title rows ─────────────────────────────────────────────────────
        ws.merge_cells("A1:N1")
        ws["A1"] = f"Daily Volume Tracker Report - {meta['report_date']}"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True)

        ws.merge_cells("A2:N2")
        ws["A2"] = (
            f"Start: {meta['month_start']} | End: {meta['month_end']} | "
            f"MTD Days: {meta['mtd_days']} | Balance: {meta['balance_days']} | "
            f"Days in Month: {meta['days_in_month']}"
        )
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)

        ws.merge_cells("A3:N3")
        ws["A3"] = meta.get("summary", "")
        ws["A3"].font = Font(name="Calibri", size=9, italic=True)

        # ── Header row ─────────────────────────────────────────────────────
        headers = [
            "Sr. No", "Plant Name",
            f"Produced Qty\n{meta.get('yesterday', '')[-5:].replace('-', '-')}",
            f"Invoiced Qty\n{meta.get('yesterday', '')[-5:].replace('-', '-')}",
            "Daily Avg\nVolume", "Req. Vol/day\nto achieve budget",
            "MTD Volume", "Extrapolated\nVol", "Budget/\nTarget",
            "% Extrapolation\nV/S Budget", "Last Month\nVolume",
            "% Variation\nLast Month", f"Last Year\n{meta.get('yesterday', '')[:7]}-Vol",
            "% Variation\nLast Year",
        ]

        row_idx = 5
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ── Column widths ──────────────────────────────────────────────────
        col_widths = [6, 28, 10, 10, 12, 14, 12, 14, 12, 14, 14, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Data rows ──────────────────────────────────────────────────────
        row_idx = 6
        sr_no = 1

        for region in report_data["regions"]:
            for plant in region["plants"]:
                row_data = [
                    sr_no, plant["plant_name"],
                    plant["daily_volume"], plant["invoiced_qty"],
                    plant["daily_avg"], plant["req_vol_day"],
                    plant["mtd_volume"], plant["extrapolated"], plant["target"],
                    plant["pct_extrap_vs_budget"], plant["last_month"],
                    plant["pct_vs_last_month"], plant["last_year"],
                    plant["pct_vs_last_year"],
                ]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font
                    cell.alignment = name_align if col_idx == 2 else data_align
                    cell.border = thin_border
                row_idx += 1
                sr_no += 1

            # Region subtotal row — skip for Unassigned
            sub = region["subtotal"]
            if sub is None:
                continue
            sub_data = [
                sub["label"], sub["region_name"],
                sub["daily_volume"], sub["invoiced_qty"],
                sub["daily_avg"], sub["req_vol_day"],
                sub["mtd_volume"], sub["extrapolated"], sub["target"],
                sub["pct_extrap_vs_budget"], sub["last_month"],
                sub["pct_vs_last_month"], sub["last_year"],
                sub["pct_vs_last_year"],
            ]
            for col_idx, val in enumerate(sub_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = region_font
                cell.fill = region_fill
                cell.alignment = name_align if col_idx == 2 else data_align
                cell.border = thin_border
            row_idx += 1

        # ── Company total row ──────────────────────────────────────────────
        ct = report_data["company_total"]
        total_data = [
            ct["label"], ct["region_name"],
            ct["daily_volume"], ct["invoiced_qty"],
            ct["daily_avg"], ct["req_vol_day"],
            ct["mtd_volume"], ct["extrapolated"], ct["target"],
            ct["pct_extrap_vs_budget"], ct["last_month"],
            ct["pct_vs_last_month"], ct["last_year"],
            ct["pct_vs_last_year"],
        ]
        for col_idx, val in enumerate(total_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = name_align if col_idx == 2 else data_align
            cell.border = thin_border

        # Freeze top rows
        ws.freeze_panes = "A6"

        wb.save(output)
        output.seek(0)
        return output

    except Exception as exc:
        logger.error(f"Excel export error: {exc}")
        raise
