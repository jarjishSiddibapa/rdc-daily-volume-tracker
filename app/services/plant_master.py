"""ERP plant-master reconciliation and active-list workbook support."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app import db
from app.models import Plant


WORKSHEET_NAME = "Updated Plant Name"
EXPECTED_HEADERS = (
    "S. No",
    "ERP Name (key)",
    "Org Code",
    "Display Name (Tracker Name)",
)


@dataclass(frozen=True)
class ActivePlantRecord:
    plant_code: str
    erp_name: str
    tracker_name: str
    row_number: int


def normalize_plant_code(value) -> str:
    """Normalize an ERP/application plant identifier for exact matching."""
    return str(value or "").strip().upper()


def load_active_plant_workbook(path: str | Path) -> dict[str, ActivePlantRecord]:
    """Load and strictly validate the approved active-plant workbook."""
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise ValueError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{WORKSHEET_NAME}' was not found")

        worksheet = workbook[WORKSHEET_NAME]
        headers = tuple(str(cell.value or "").strip() for cell in worksheet[1][:4])
        workbook_rows = [
            (
                row[0].row,
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
            )
            for row in worksheet.iter_rows(min_row=2, max_col=4)
            if any(cell.value is not None for cell in row)
        ]
    finally:
        workbook.close()

    if headers != EXPECTED_HEADERS:
        raise ValueError(
            "Unexpected workbook columns. Expected: " + ", ".join(EXPECTED_HEADERS)
        )

    records: dict[str, ActivePlantRecord] = {}
    for row_number, _serial, erp_value, code_value, tracker_value in workbook_rows:
        erp_name = str(erp_value or "").strip()
        plant_code = normalize_plant_code(code_value)
        tracker_name = str(tracker_value or "").strip()

        if not plant_code or not erp_name or not tracker_name:
            raise ValueError(
                f"Row {row_number} must contain ERP name, organization code, and tracker name"
            )
        if len(plant_code) > 50 or len(erp_name) > 100 or len(tracker_name) > 100:
            raise ValueError(f"Row {row_number} exceeds an application field length")
        if plant_code in records:
            raise ValueError(
                f"Duplicate organization code {plant_code} at rows "
                f"{records[plant_code].row_number} and {row_number}"
            )

        records[plant_code] = ActivePlantRecord(
            plant_code=plant_code,
            erp_name=erp_name,
            tracker_name=tracker_name,
            row_number=row_number,
        )

    if not records:
        raise ValueError("The active-plant workbook contains no plant rows")
    return records


def normalize_erp_organizations(
    organizations: Iterable[dict],
) -> dict[str, str]:
    """Normalize Oracle organization rows into a unique code-to-name mapping."""
    normalized: dict[str, str] = {}
    for record in organizations:
        code = normalize_plant_code(record.get("organization_code"))
        name = str(record.get("organization_name") or "").strip()
        if not code:
            continue
        if code in normalized and normalized[code] != name:
            raise ValueError(f"ERP returned conflicting names for organization {code}")
        normalized[code] = name
    return normalized


def validate_active_plants_against_erp(
    active_plants: dict[str, ActivePlantRecord],
    erp_organizations: dict[str, str],
) -> list[dict]:
    """Validate code membership and report non-blocking ERP-name differences."""
    missing_codes = sorted(set(active_plants) - set(erp_organizations))
    if missing_codes:
        raise ValueError(
            "Active organization codes missing from ERP: " + ", ".join(missing_codes)
        )

    return [
        {
            "plant_code": code,
            "workbook_erp_name": record.erp_name,
            "current_erp_name": erp_organizations[code],
        }
        for code, record in sorted(active_plants.items())
        if record.erp_name.casefold() != erp_organizations[code].casefold()
    ]


def reconcile_plant_master(
    active_plants: dict[str, ActivePlantRecord],
    erp_organizations: dict[str, str],
    *,
    apply_changes: bool,
) -> dict:
    """Build or apply the approved active/inactive plant baseline.

    Every current application plant not listed in ``active_plants`` is made
    inactive. Every enabled ERP organization is materialized in MySQL so it is
    remembered as existing; organizations outside the approved list start
    inactive. This lets later ERP syncs distinguish a genuinely new code and
    default only that new code to active.
    """
    existing_plants = {
        normalize_plant_code(plant.plant_code): plant for plant in Plant.query.all()
    }

    activated = []
    deactivated = []
    tracker_names_updated = []
    erp_names_updated = []
    created_active = []
    created_inactive = []

    for code, plant in existing_plants.items():
        should_be_active = code in active_plants
        if bool(plant.is_active) != should_be_active:
            (activated if should_be_active else deactivated).append(code)
            if apply_changes:
                plant.is_active = should_be_active

        if code in active_plants:
            tracker_name = active_plants[code].tracker_name
            if (plant.daily_tracker_name or "").strip() != tracker_name:
                tracker_names_updated.append(code)
                if apply_changes:
                    plant.daily_tracker_name = tracker_name

        erp_name = erp_organizations.get(code)
        if erp_name and (plant.erp_name or "").strip() != erp_name:
            erp_names_updated.append(code)
            if apply_changes:
                plant.erp_name = erp_name

    for code, erp_name in sorted(erp_organizations.items()):
        if code in existing_plants:
            continue

        is_active = code in active_plants
        tracker_name = (
            active_plants[code].tracker_name if is_active else (erp_name or code)
        )
        (created_active if is_active else created_inactive).append(code)
        if apply_changes:
            db.session.add(
                Plant(
                    plant_code=code,
                    daily_tracker_name=tracker_name,
                    erp_name=erp_name,
                    region="",
                    is_active=is_active,
                    is_manual_entry=False,
                )
            )

    return {
        "approved_active": len(active_plants),
        "erp_organizations": len(erp_organizations),
        "existing_plants": len(existing_plants),
        "activated": activated,
        "deactivated": deactivated,
        "tracker_names_updated": tracker_names_updated,
        "erp_names_updated": erp_names_updated,
        "created_active": created_active,
        "created_inactive": created_inactive,
        "apply_changes": apply_changes,
    }
