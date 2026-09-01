"""Monthly target management routes."""

from datetime import date, datetime

from flask import Blueprint, render_template, jsonify, request, send_file
from flask_login import login_required

from app import db
from app.models import Plant, PlantMonthlyTarget
from app.services.excel_service import parse_target_excel
from app.decorators import admin_required, manual_entry_required, targets_required
from app.services.audit import log_action
from app.services.report_generator import invalidate_report_cache

targets_bp = Blueprint("targets", __name__)


def _combined_name(plant) -> str:
    t = (plant.daily_tracker_name or "").strip()
    e = (plant.erp_name or "").strip()
    if t and e and t != e:
        return f"{t} ({e})"
    return t or e or plant.plant_code


@targets_bp.route("/targets")
@login_required
def targets_page():
    """Target management page. Readable by all logged-in users."""
    return render_template("targets.html")


@targets_bp.route("/api/targets/<month_str>")
@login_required
def api_get_targets(month_str):
    """
    Get targets for a given month, grouped by region.
    month_str format: YYYY-MM
    """
    try:
        year, month = month_str.split("-")
        month_date = date(int(year), int(month), 1)
    except (ValueError, IndexError):
        return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

    from app.services.report_generator import get_region_order

    plants = Plant.query.filter_by(is_active=True).order_by(Plant.region, Plant.display_order, Plant.daily_tracker_name).all()

    targets = {
        t.plant_code: float(t.target_volume)
        for t in PlantMonthlyTarget.query.filter_by(month_date=month_date).all()
    }

    # Auto-carry forward from previous month if no targets exist
    if not targets:
        if month_date.month == 1:
            prev_month = date(month_date.year - 1, 12, 1)
        else:
            prev_month = date(month_date.year, month_date.month - 1, 1)
        prev_targets = PlantMonthlyTarget.query.filter_by(month_date=prev_month).all()
        if prev_targets:
            for pt in prev_targets:
                new_t = PlantMonthlyTarget(
                    plant_code=pt.plant_code,
                    month_date=month_date,
                    target_volume=pt.target_volume,
                )
                db.session.add(new_t)
            try:
                db.session.commit()
                invalidate_report_cache()
            except Exception:
                db.session.rollback()
                # Concurrent request may have won the race — re-fetch instead of returning empty
            targets = {
                t.plant_code: float(t.target_volume)
                for t in PlantMonthlyTarget.query.filter_by(month_date=month_date).all()
            }

    # Group by region
    region_map = {}
    for plant in plants:
        r = plant.region or "Other"
        if r not in region_map:
            region_map[r] = []
        region_map[r].append({
            "plant_code": plant.plant_code,
            "plant_name": _combined_name(plant),
            "region": r,
            "target_volume": targets.get(plant.plant_code, 0.0),
        })

    # Order regions same as dashboard (live DB order)
    ordered_regions = []
    for rname in get_region_order():
        if rname in region_map:
            ordered_regions.append({"region": rname, "plants": region_map.pop(rname)})
    for rname, plist in region_map.items():
        ordered_regions.append({"region": rname, "plants": plist})

    # Flat list for backward compat
    flat = []
    for rg in ordered_regions:
        flat.extend(rg["plants"])

    return jsonify({
        "month": month_date.isoformat(),
        "targets": flat,
        "regions": ordered_regions,
    })


@targets_bp.route("/api/targets/<month_str>/template")
@login_required
def api_download_template(month_str):
    """Download Excel template with plant codes/names and current target values."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    try:
        year, month = month_str.split("-")
        month_date = date(int(year), int(month), 1)
    except (ValueError, IndexError):
        return jsonify({"error": "Invalid month format"}), 400

    from app.services.report_generator import get_region_order

    plants = Plant.query.filter_by(is_active=True).order_by(Plant.region, Plant.display_order, Plant.daily_tracker_name).all()
    targets = {
        t.plant_code: float(t.target_volume)
        for t in PlantMonthlyTarget.query.filter_by(month_date=month_date).all()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Targets"

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    region_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    region_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    headers = ["plant_code", "plant_name", "region", "target_volume"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Group by region
    region_map = {}
    for p in plants:
        r = p.region or "Other"
        region_map.setdefault(r, []).append(p)

    ordered = []
    for rname in get_region_order():
        if rname in region_map:
            ordered.append((rname, region_map.pop(rname)))
    for rname, plist in region_map.items():
        ordered.append((rname, plist))

    row = 2
    for rname, plist in ordered:
        # Region header
        cell = ws.cell(row=row, column=1, value=rname)
        cell.font = region_font
        cell.fill = region_fill
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = region_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        for p in plist:
            ws.cell(row=row, column=1, value=p.plant_code).border = thin_border
            ws.cell(row=row, column=2, value=_combined_name(p)).border = thin_border
            ws.cell(row=row, column=3, value=p.region or "").border = thin_border
            tv = targets.get(p.plant_code, 0)
            c4 = ws.cell(row=row, column=4, value=tv)
            c4.border = thin_border
            c4.number_format = '#,##0'
            row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Target_Template_{month_str}.xlsx"
    log_action("target_template_downloaded", {"month": month_str})
    db.session.commit()

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@targets_bp.route("/api/targets", methods=["PUT"])
@targets_required
def api_update_targets():
    """
    Update targets inline. Admin only.

    Body: { month: "YYYY-MM", targets: [{ plant_code, target_volume }, ...] }
    """
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    month_str = data.get("month", "")
    try:
        year, month = month_str.split("-")
        month_date = date(int(year), int(month), 1)
    except (ValueError, IndexError):
        return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

    target_entries = data.get("targets", [])
    saved = 0
    errors = []

    for entry in target_entries:
        plant_code = entry.get("plant_code", "").strip()
        try:
            target_value = float(entry.get("target_volume", 0))
        except (ValueError, TypeError):
            errors.append(f"{plant_code}: invalid target value")
            continue

        if target_value < 0:
            errors.append(f"{plant_code}: target volume cannot be negative")
            continue

        plant = db.session.get(Plant, plant_code)
        if not plant:
            errors.append(f"{plant_code}: plant not found")
            continue

        existing = PlantMonthlyTarget.query.filter_by(
            plant_code=plant_code, month_date=month_date
        ).first()

        if existing:
            existing.target_volume = target_value
        else:
            target = PlantMonthlyTarget(
                plant_code=plant_code,
                month_date=month_date,
                target_volume=target_value,
            )
            db.session.add(target)
        saved += 1

    try:
        db.session.commit()
        invalidate_report_cache()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Database error: {str(exc)}"}), 500

    log_action("target_update", {"month": month_str, "saved": saved})
    db.session.commit()

    return jsonify({"status": "success", "saved": saved, "errors": errors})


@targets_bp.route("/api/targets/upload", methods=["POST"])
@targets_required
def api_upload_targets():
    """Upload an Excel file with monthly targets. Admin only."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    month_str = request.form.get("month", "")
    try:
        year, month = month_str.split("-")
        month_date = date(int(year), int(month), 1)
    except (ValueError, IndexError):
        return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

    result = parse_target_excel(file.stream, month_date)
    status_code = 200 if result["status"] == "success" else 400

    if result["status"] == "success":
        invalidate_report_cache()
        log_action("target_upload", {"month": month_str, "file": file.filename})
        db.session.commit()

    return jsonify(result), status_code
