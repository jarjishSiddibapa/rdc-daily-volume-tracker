"""Plant management CRUD routes — admin only."""

from flask import Blueprint, render_template, jsonify, request, send_file

from app import db
from app.models import Plant
from app.decorators import admin_required
from app.services.audit import log_action
from app.services.report_generator import invalidate_report_cache

plants_bp = Blueprint("plants", __name__)


@plants_bp.route("/plants")
@admin_required
def plants_page():
    """Plant management page. Admin only."""
    return render_template("plants.html")


@plants_bp.route("/api/plants")
@admin_required
def api_list_plants():
    """List all plants with optional region filter. Admin only."""
    region = request.args.get("region")
    query = Plant.query.order_by(Plant.region, Plant.display_order, Plant.daily_tracker_name)

    if region:
        query = query.filter_by(region=region)

    plants = query.all()
    regions = sorted(set(p.region for p in Plant.query.with_entities(Plant.region).distinct().all() if p.region))

    return jsonify({
        "plants": [p.to_dict() for p in plants],
        "regions": regions,
    })


@plants_bp.route("/api/plants", methods=["POST"])
@admin_required
def api_add_plant():
    """Add a new plant. Admin only."""
    data = request.get_json()

    if not data or not data.get("plant_code"):
        return jsonify({"error": "plant_code is required"}), 400

    plant_code = data["plant_code"].strip().upper()

    if db.session.get(Plant, plant_code):
        return jsonify({"error": f"Plant {plant_code} already exists"}), 409

    plant = Plant(
        plant_code=plant_code,
        daily_tracker_name=data.get("daily_tracker_name", "").strip(),
        erp_name=data.get("erp_name", "").strip(),
        region=data.get("region", "").strip(),
        is_active=data.get("is_active", True),
    )

    db.session.add(plant)

    try:
        db.session.commit()
        invalidate_report_cache()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500

    log_action("plant_create", {"plant_code": plant_code, "name": plant.daily_tracker_name})
    db.session.commit()

    return jsonify({"status": "success", "plant": plant.to_dict()}), 201


@plants_bp.route("/api/plants/<plant_code>", methods=["PUT"])
@admin_required
def api_update_plant(plant_code):
    """Update an existing plant. Admin only."""
    plant = db.session.get(Plant, plant_code)
    if not plant:
        return jsonify({"error": "Plant not found"}), 404

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    # Capture originals for audit detail
    old_active = plant.is_active
    old_region = plant.region
    old_name = plant.daily_tracker_name

    if "daily_tracker_name" in data:
        plant.daily_tracker_name = data["daily_tracker_name"].strip()
    if "erp_name" in data:
        plant.erp_name = data["erp_name"].strip()
    if "region" in data:
        plant.region = data["region"].strip()
    if "is_active" in data:
        plant.is_active = bool(data["is_active"])

    try:
        db.session.commit()
        invalidate_report_cache()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500

    log_action("plant_update", {
        "plant_code": plant_code,
        "name": plant.daily_tracker_name,
        "active_changed": old_active != plant.is_active,
        "is_active": plant.is_active,
        "region_changed": old_region != plant.region,
        "old_region": old_region,
        "new_region": plant.region,
    })
    db.session.commit()

    return jsonify({"status": "success", "plant": plant.to_dict()})


@plants_bp.route("/api/plants/<plant_code>", methods=["DELETE"])
@admin_required
def api_delete_plant(plant_code):
    """Delete a plant (cascades to volumes and targets). Admin only."""
    plant = db.session.get(Plant, plant_code)
    if not plant:
        return jsonify({"error": "Plant not found"}), 404

    try:
        db.session.delete(plant)
        db.session.commit()
        invalidate_report_cache()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500

    log_action("plant_delete", {"plant_code": plant_code})
    db.session.commit()

    return jsonify({"status": "success", "message": f"Plant {plant_code} deleted"})



@plants_bp.route("/api/plants/reorder", methods=["PUT"])
@admin_required
def api_reorder_plants():
    """Update display_order for plants within a region."""
    data = request.get_json()
    if not data or "region" not in data or "order" not in data:
        return jsonify({"error": "region and order required"}), 400

    region = data["region"]
    order = data["order"]   # list of plant_codes in desired order

    for i, plant_code in enumerate(order):
        plant = db.session.get(Plant, plant_code)
        if plant and plant.region == region:
            plant.display_order = i + 1

    try:
        db.session.commit()
        invalidate_report_cache()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500

    log_action("plant_reorder", {"region": region, "order": order})
    db.session.commit()
    return jsonify({"status": "success"})


@plants_bp.route("/api/plants/download")
@admin_required
def api_download_plants():
    """Download all plants as an Excel file. Admin only."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    from app.services.report_generator import REGION_ORDER

    plants = Plant.query.order_by(Plant.region, Plant.display_order, Plant.daily_tracker_name).all()

    # Group by region
    region_map = {}
    for p in plants:
        r = p.region or "Other"
        region_map.setdefault(r, []).append(p)

    ordered = []
    for rname in REGION_ORDER:
        if rname in region_map:
            ordered.append((rname, region_map.pop(rname)))
    for rname, plist in region_map.items():
        ordered.append((rname, plist))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plants"

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    region_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    region_font = Font(bold=True, size=11)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Plant Code", "Tracker Name", "ERP Name", "Region", "Active", "Manual Entry"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    row = 2
    for rname, plist in ordered:
        # Region header
        cell = ws.cell(row=row, column=1, value=rname)
        cell.font = region_font
        cell.fill = region_fill
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = region_fill
            ws.cell(row=row, column=c).border = thin
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        for p in plist:
            ws.cell(row=row, column=1, value=p.plant_code).border = thin
            ws.cell(row=row, column=2, value=p.daily_tracker_name or "").border = thin
            ws.cell(row=row, column=3, value=p.erp_name or "").border = thin
            ws.cell(row=row, column=4, value=p.region or "").border = thin
            c5 = ws.cell(row=row, column=5, value="Yes" if p.is_active else "No")
            c5.border = thin
            c5.alignment = Alignment(horizontal="center")
            c6 = ws.cell(row=row, column=6, value="Yes" if p.is_manual_entry else "No")
            c6.border = thin
            c6.alignment = Alignment(horizontal="center")
            row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_action("plant_list_downloaded", {"plant_count": len(plants)})
    db.session.commit()

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Plants_Data.xlsx",
    )
