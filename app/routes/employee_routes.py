"""Employee & TM Details routes."""

import io
import logging
from flask import Blueprint, render_template, jsonify, request, send_file
from flask_login import current_user

from app import db
from app.models import Plant, PlantEmployeeDetails
from app.decorators import employee_details_required
from app.services.audit import log_action

employee_bp = Blueprint("employee", __name__)
logger = logging.getLogger(__name__)


@employee_bp.route("/employee-details")
@employee_details_required
def employee_details_page():
    return render_template("employee_details.html")


@employee_bp.route("/api/employee-details")
@employee_details_required
def api_list_employee_details():
    """Return all active plants with their employee details."""
    plants = (
        Plant.query
        .filter_by(is_active=True)
        .order_by(Plant.region, Plant.daily_tracker_name)
        .all()
    )

    # Pre-fetch all employee detail rows
    details_map = {
        d.plant_code: d
        for d in PlantEmployeeDetails.query.all()
    }

    result = []
    for p in plants:
        d = details_map.get(p.plant_code)
        result.append({
            "plant_code": p.plant_code,
            "plant_name": (p.daily_tracker_name or p.erp_name or p.plant_code).strip(),
            "region": p.region or "Unassigned",
            "on_roll":   d.on_roll   if d else 0,
            "teamlease": d.teamlease if d else 0,
            "no_of_tm":  d.no_of_tm  if d else 0,
            "updated_at": d.updated_at.strftime("%d-%m-%Y %H:%M") if d and d.updated_at else None,
            "updated_by": d.updated_by or "" if d else "",
        })

    return jsonify({
        "plants": result,
        "can_edit": current_user.role == "admin" or current_user.can_edit_employee_details,
    })


@employee_bp.route("/api/employee-details/<plant_code>", methods=["PUT"])
@employee_details_required
def api_update_employee_details(plant_code):
    """Upsert employee details for a plant."""
    plant = db.session.get(Plant, plant_code)
    if not plant or not plant.is_active:
        return jsonify({"error": "Plant not found"}), 404

    data = request.get_json() or {}

    try:
        on_roll   = int(data.get("on_roll", 0))
        teamlease = int(data.get("teamlease", 0))
        no_of_tm  = int(data.get("no_of_tm", 0))
    except (ValueError, TypeError):
        return jsonify({"error": "Values must be integers"}), 400

    if on_roll < 0 or teamlease < 0 or no_of_tm < 0:
        return jsonify({"error": "Values cannot be negative"}), 400

    existing = db.session.get(PlantEmployeeDetails, plant_code)
    if existing:
        existing.on_roll   = on_roll
        existing.teamlease = teamlease
        existing.no_of_tm  = no_of_tm
        existing.updated_by = current_user.username
    else:
        db.session.add(PlantEmployeeDetails(
            plant_code=plant_code,
            on_roll=on_roll,
            teamlease=teamlease,
            no_of_tm=no_of_tm,
            updated_by=current_user.username,
        ))

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Employee details save failed for {plant_code}: {exc}")
        return jsonify({"error": "Database error — please try again."}), 500

    log_action("employee_details_updated", {
        "plant_code": plant_code,
        "plant_name": plant.daily_tracker_name or plant.erp_name,
        "on_roll": on_roll,
        "teamlease": teamlease,
        "no_of_tm": no_of_tm,
    })
    db.session.commit()

    return jsonify({"success": True, "message": f"Details saved for {plant_code}"})


# ── Bulk Excel download ───────────────────────────────────────────────────────

@employee_bp.route("/api/employee-details/download")
@employee_details_required
def api_download_employee_template():
    """Generate and return a pre-filled Excel template for bulk editing."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, Protection
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    plants = (
        Plant.query
        .filter_by(is_active=True)
        .order_by(Plant.region, Plant.daily_tracker_name)
        .all()
    )
    details_map = {
        d.plant_code: d
        for d in PlantEmployeeDetails.query.all()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee & TM Details"

    # ── Styles ───────────────────────────────────────────────────────────────
    # Identifier columns (read-only feel) — light grey
    id_fill   = PatternFill("solid", fgColor="D9D9D9")
    # Editable columns — light yellow
    edit_fill = PatternFill("solid", fgColor="FFF9C4")
    # Header row — dark blue
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)

    body_font      = Font(size=10)
    id_font        = Font(size=10, color="444444")
    center_align   = Alignment(horizontal="center", vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center")
    thin           = Side(border_style="thin", color="CCCCCC")
    cell_border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Instruction row ───────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    inst = ws["A1"]
    inst.value = (
        "INSTRUCTIONS: Fill / update the On Roll, Teamlease, and No. of TMs columns only. "
        "Do NOT modify Plant Code, ERP Name, or Tracker Name — they are used to match rows."
    )
    inst.font      = Font(italic=True, color="7F0000", size=9)
    inst.fill      = PatternFill("solid", fgColor="FFF3CD")
    inst.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header row (row 2) ────────────────────────────────────────────────────
    HEADERS = [
        "Plant Code",   # A — identifier
        "ERP Name",     # B — identifier
        "Tracker Name", # C — identifier
        "Area",         # D — identifier
        "On Roll",      # E — editable
        "Teamlease",    # F — editable
        "No. of TMs",   # G — editable
    ]
    ID_COLS   = {1, 2, 3, 4}   # 1-based column numbers
    EDIT_COLS = {5, 6, 7}

    for col_idx, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center_align
        cell.border    = cell_border
    ws.row_dimensions[2].height = 20

    # ── Data rows (starting row 3) ────────────────────────────────────────────
    for row_idx, p in enumerate(plants, start=3):
        d = details_map.get(p.plant_code)
        tracker = (p.daily_tracker_name or "").strip()
        erp     = (p.erp_name or "").strip()
        area    = (p.region or "Unassigned").strip()

        row_data = [
            p.plant_code,
            erp,
            tracker,
            area,
            d.on_roll   if d else 0,
            d.teamlease if d else 0,
            d.no_of_tm  if d else 0,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = cell_border
            cell.alignment = center_align if col_idx in EDIT_COLS | {1} else left_align

            if col_idx in ID_COLS:
                cell.fill = id_fill
                cell.font = id_font
            else:
                cell.fill = edit_fill
                cell.font = body_font

    # ── Column widths ─────────────────────────────────────────────────────────
    COL_WIDTHS = [14, 28, 28, 18, 12, 12, 12]
    for idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── Freeze panes — keep header rows pinned when scrolling ────────────────
    ws.freeze_panes = "A3"

    # ── Numeric data-validation on editable columns ───────────────────────────
    if len(plants) > 0:
        last_row = 2 + len(plants)
        for col_letter in ("E", "F", "G"):
            dv = DataValidation(
                type="whole",
                operator="greaterThanOrEqual",
                formula1="0",
                showErrorMessage=True,
                error="Enter a whole number >= 0",
                errorTitle="Invalid value",
            )
            dv.sqref = f"{col_letter}3:{col_letter}{last_row}"
            ws.add_data_validation(dv)

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="employee_tm_details.xlsx",
    )


# ── Bulk Excel upload ─────────────────────────────────────────────────────────

@employee_bp.route("/api/employee-details/upload", methods=["POST"])
@employee_details_required
def api_upload_employee_details():
    """Parse an uploaded Excel file and bulk-update employee details."""
    from openpyxl import load_workbook

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "No file uploaded"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx / .xlsm files are accepted"}), 400

    _MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
    raw = file.read(_MAX_UPLOAD_BYTES + 1)
    if len(raw) > _MAX_UPLOAD_BYTES:
        return jsonify({"error": "File is too large (max 5 MB)"}), 400

    try:
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception:
        return jsonify({"error": "Could not read the file. Make sure it is a valid Excel workbook."}), 400

    ws = wb.active

    # ── Find the header row ───────────────────────────────────────────────────
    REQUIRED_HEADERS = {"plant code", "on roll", "teamlease", "no. of tms"}
    header_row_idx   = None
    col_map          = {}  # normalised header → 0-based column index

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalised = {
            str(cell).strip().lower(): col_i
            for col_i, cell in enumerate(row)
            if cell is not None
        }
        if REQUIRED_HEADERS.issubset(normalised.keys()):
            header_row_idx = row_idx
            col_map        = normalised
            break

    if header_row_idx is None:
        return jsonify({
            "error": "Could not find required header row. "
                     "Make sure the file contains columns: Plant Code, On Roll, Teamlease, No. of TMs."
        }), 400

    # Column indices (0-based)
    ci_code      = col_map["plant code"]
    ci_on_roll   = col_map["on roll"]
    ci_teamlease = col_map["teamlease"]
    ci_tm        = col_map["no. of tms"]

    # ── Load all active plant codes for validation ────────────────────────────
    active_codes = {
        p.plant_code
        for p in Plant.query.filter_by(is_active=True).all()
    }

    # ── Process data rows ─────────────────────────────────────────────────────
    updated  = 0
    skipped  = []   # (plant_code, reason)
    errors   = []

    for row in ws.iter_rows(
        min_row=header_row_idx + 1,
        values_only=True,
    ):
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue

        raw_code = row[ci_code]
        if raw_code is None:
            continue
        plant_code = str(raw_code).strip()

        if plant_code not in active_codes:
            skipped.append({"plant_code": plant_code, "reason": "Not found or inactive"})
            continue

        # Parse numeric values — treat None / blank as 0
        def _int(val):
            if val is None or str(val).strip() == "":
                return 0
            try:
                return max(0, int(float(str(val).strip())))
            except (ValueError, TypeError):
                return None

        on_roll   = _int(row[ci_on_roll])
        teamlease = _int(row[ci_teamlease])
        no_of_tm  = _int(row[ci_tm])

        if on_roll is None or teamlease is None or no_of_tm is None:
            errors.append({"plant_code": plant_code, "reason": "Non-numeric value in On Roll / Teamlease / No. of TMs"})
            continue

        # Upsert
        existing = db.session.get(PlantEmployeeDetails, plant_code)
        if existing:
            existing.on_roll   = on_roll
            existing.teamlease = teamlease
            existing.no_of_tm  = no_of_tm
            existing.updated_by = current_user.username
        else:
            db.session.add(PlantEmployeeDetails(
                plant_code=plant_code,
                on_roll=on_roll,
                teamlease=teamlease,
                no_of_tm=no_of_tm,
                updated_by=current_user.username,
            ))
        updated += 1

    if updated == 0 and errors:
        db.session.rollback()
        return jsonify({"error": "No rows could be updated.", "errors": errors}), 400

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Database error: {exc}"}), 500

    log_action("employee_details_bulk_upload", {
        "updated":      updated,
        "skipped":      len(skipped),
        "errors":       len(errors),
        "uploaded_by":  current_user.username,
    })
    db.session.commit()

    return jsonify({
        "success":  True,
        "updated":  updated,
        "skipped":  skipped,
        "errors":   errors,
        "message":  f"{updated} plant(s) updated successfully."
                    + (f" {len(skipped)} skipped." if skipped else "")
                    + (f" {len(errors)} error(s)." if errors else ""),
    })
