"""
Analytics service — dynamic day-wise and monthly production volume reports.

Day-wise : per-plant daily volumes over a date range  (PlantDailyVolume).
Monthly  : per-plant monthly totals over a month range (PlantMonthlyVolume,
           falling back to daily-sum for the current still-open month).
"""

import io
import logging
from datetime import date, timedelta

from sqlalchemy import func

from app import db
from app.models import Plant, PlantDailyVolume, PlantMonthlyVolume
from app.services.report_generator import get_region_order, _combined_name, _safe_float

logger = logging.getLogger(__name__)

_MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


# ── iteration helpers ─────────────────────────────────────────────────────────

def _iter_dates(from_date: date, to_date: date):
    d = from_date
    while d <= to_date:
        yield d
        d += timedelta(days=1)


def _iter_months(from_month: date, to_month: date):
    m = date(from_month.year, from_month.month, 1)
    end = date(to_month.year, to_month.month, 1)
    while m <= end:
        yield m
        m = date(m.year + (1 if m.month == 12 else 0), 1 if m.month == 12 else m.month + 1, 1)


def _group_and_sort(plant_rows: list) -> list:
    """Group plant rows by region, compute per-column subtotals, sort by region order."""
    n_cols = len(plant_rows[0]["volumes"]) if plant_rows else 0
    has_inv = plant_rows and "invoiced_volumes" in plant_rows[0]

    region_map: dict = {}
    for row in plant_rows:
        rg = row["region"]
        region_map.setdefault(rg, {"name": rg, "plants": []})["plants"].append(row)

    for rg, rdata in region_map.items():
        plist = rdata["plants"]
        sub_vols = [round(sum(p["volumes"][i] for p in plist), 2) for i in range(n_cols)]
        rdata["subtotal"] = {
            "volumes": sub_vols,
            "total": round(sum(sub_vols), 2),
        }
        if has_inv:
            sub_inv = [round(sum(p["invoiced_volumes"][i] for p in plist), 2) for i in range(n_cols)]
            rdata["subtotal"]["invoiced_volumes"] = sub_inv
            rdata["subtotal"]["invoiced_total"] = round(sum(sub_inv), 2)

    order = get_region_order()
    ordered, rem = [], dict(region_map)
    for rn in order:
        if rn in rem:
            ordered.append(rem.pop(rn))
    for rn, rd in list(rem.items()):
        if rn != "Unassigned":
            ordered.append(rd)
    if "Unassigned" in rem:
        ordered.append(rem["Unassigned"])
    return ordered


# ── day-wise report ───────────────────────────────────────────────────────────

def generate_daywise_report(
    from_date: date,
    to_date: date,
    hide_zero: bool = False,
    plant_codes: list = None,
) -> dict:
    """Return per-plant daily volumes for [from_date … to_date] inclusive.

    If plant_codes is provided, only those plants are included.
    """
    dates = list(_iter_dates(from_date, to_date))
    if not dates:
        return {"error": "Empty date range"}

    rows = (
        db.session.query(
            PlantDailyVolume.plant_code,
            PlantDailyVolume.entry_date,
            PlantDailyVolume.volume,
            PlantDailyVolume.invoiced_qty,
        )
        .filter(
            PlantDailyVolume.entry_date >= from_date,
            PlantDailyVolume.entry_date <= to_date,
        )
        .all()
    )

    vol_map: dict = {}
    inv_map: dict = {}
    for r in rows:
        vol_map.setdefault(r.plant_code, {})[r.entry_date] = _safe_float(r.volume)
        inv_map.setdefault(r.plant_code, {})[r.entry_date] = _safe_float(r.invoiced_qty)

    plant_q = Plant.query.filter_by(is_active=True).order_by(
        Plant.region, Plant.display_order, Plant.daily_tracker_name
    )
    if plant_codes:
        plant_q = plant_q.filter(Plant.plant_code.in_(plant_codes))
    plants = plant_q.all()

    plant_rows = []
    for plant in plants:
        pc = plant.plant_code
        vols = [vol_map.get(pc, {}).get(d, 0.0) for d in dates]
        invs = [inv_map.get(pc, {}).get(d, 0.0) for d in dates]
        total = sum(vols)
        if hide_zero and total == 0:
            continue
        plant_rows.append(
            {
                "plant_code": pc,
                "daily_tracker_name": plant.daily_tracker_name or "",
                "erp_name": plant.erp_name or "",
                "plant_name": _combined_name(plant),
                "region": plant.region or "Unassigned",
                "is_manual_entry": plant.is_manual_entry,
                "volumes": [round(v, 2) for v in vols],
                "total": round(total, 2),
                "invoiced_volumes": [round(v, 2) for v in invs],
                "invoiced_total": round(sum(invs), 2),
            }
        )

    n = len(dates)
    empty_result = {
        "type": "daywise",
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "dates": [d.isoformat() for d in dates],
        "date_labels": [f"{d.day}-{_MONTH_NAMES[d.month - 1]}" for d in dates],
        "regions": [],
        "company_total": {"volumes": [0.0] * n, "total": 0.0,
                          "invoiced_volumes": [0.0] * n, "invoiced_total": 0.0},
    }
    if not plant_rows:
        return empty_result

    ordered = _group_and_sort(plant_rows)
    comp_vols = [round(sum(p["volumes"][i] for p in plant_rows), 2) for i in range(n)]
    comp_inv  = [round(sum(p["invoiced_volumes"][i] for p in plant_rows), 2) for i in range(n)]

    return {
        "type": "daywise",
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "dates": [d.isoformat() for d in dates],
        "date_labels": [f"{d.day}-{_MONTH_NAMES[d.month - 1]}" for d in dates],
        "regions": ordered,
        "company_total": {
            "volumes": comp_vols,
            "total": round(sum(comp_vols), 2),
            "invoiced_volumes": comp_inv,
            "invoiced_total": round(sum(comp_inv), 2),
        },
    }


# ── monthly report ────────────────────────────────────────────────────────────

def generate_monthly_report(
    from_month: date,
    to_month: date,
    hide_zero: bool = False,
    plant_codes: list = None,
) -> dict:
    """Return per-plant monthly totals for [from_month … to_month] inclusive.

    Uses PlantMonthlyVolume for closed months; sums PlantDailyVolume for the
    current still-open month when no monthly record exists yet.
    If plant_codes is provided, only those plants are included.
    """
    months = list(_iter_months(from_month, to_month))
    if not months:
        return {"error": "Empty month range"}

    rows = (
        db.session.query(
            PlantMonthlyVolume.plant_code,
            PlantMonthlyVolume.month_date,
            PlantMonthlyVolume.total_actual_volume,
        )
        .filter(
            PlantMonthlyVolume.month_date >= months[0],
            PlantMonthlyVolume.month_date <= months[-1],
        )
        .all()
    )

    vol_map: dict = {}
    for r in rows:
        vol_map.setdefault(r.plant_code, {})[r.month_date] = _safe_float(
            r.total_actual_volume
        )

    # Current open month — fall back to daily sum if no monthly record yet
    today = date.today()
    cur_month = date(today.year, today.month, 1)
    if cur_month in months:
        daily_agg = (
            db.session.query(
                PlantDailyVolume.plant_code,
                func.sum(PlantDailyVolume.volume).label("vol"),
            )
            .filter(
                PlantDailyVolume.entry_date >= cur_month,
                PlantDailyVolume.entry_date <= today,
            )
            .group_by(PlantDailyVolume.plant_code)
            .all()
        )
        for r in daily_agg:
            pc_map = vol_map.setdefault(r.plant_code, {})
            if cur_month not in pc_map:        # don't overwrite a closed-month record
                pc_map[cur_month] = _safe_float(r.vol)

    # Invoiced qty — always summed from daily records (no monthly aggregation for invoiced)
    # Compute the end of the last month in range (clamped to today)
    last_month_end = date(
        months[-1].year + (1 if months[-1].month == 12 else 0),
        (months[-1].month % 12) + 1,
        1,
    ) - timedelta(days=1)
    inv_range_end = min(last_month_end, today)

    daily_inv_rows = (
        db.session.query(
            PlantDailyVolume.plant_code,
            PlantDailyVolume.entry_date,
            PlantDailyVolume.invoiced_qty,
        )
        .filter(
            PlantDailyVolume.entry_date >= months[0],
            PlantDailyVolume.entry_date <= inv_range_end,
        )
        .all()
    )
    inv_map: dict = {}
    for r in daily_inv_rows:
        m = date(r.entry_date.year, r.entry_date.month, 1)
        pc_inv = inv_map.setdefault(r.plant_code, {})
        pc_inv[m] = pc_inv.get(m, 0.0) + _safe_float(r.invoiced_qty)

    plant_q = Plant.query.filter_by(is_active=True).order_by(
        Plant.region, Plant.display_order, Plant.daily_tracker_name
    )
    if plant_codes:
        plant_q = plant_q.filter(Plant.plant_code.in_(plant_codes))
    plants = plant_q.all()

    plant_rows = []
    for plant in plants:
        pc = plant.plant_code
        vols = [vol_map.get(pc, {}).get(m, 0.0) for m in months]
        invs = [inv_map.get(pc, {}).get(m, 0.0) for m in months]
        total = sum(vols)
        if hide_zero and total == 0:
            continue
        plant_rows.append(
            {
                "plant_code": pc,
                "daily_tracker_name": plant.daily_tracker_name or "",
                "erp_name": plant.erp_name or "",
                "plant_name": _combined_name(plant),
                "region": plant.region or "Unassigned",
                "is_manual_entry": plant.is_manual_entry,
                "volumes": [round(v, 2) for v in vols],
                "total": round(total, 2),
                "invoiced_volumes": [round(v, 2) for v in invs],
                "invoiced_total": round(sum(invs), 2),
            }
        )

    month_labels = [
        f"{_MONTH_NAMES[m.month - 1]} '{str(m.year)[2:]}" for m in months
    ]
    n = len(months)

    if not plant_rows:
        return {
            "type": "monthly",
            "from_month": from_month.isoformat(),
            "to_month": to_month.isoformat(),
            "months": [m.isoformat() for m in months],
            "month_labels": month_labels,
            "regions": [],
            "company_total": {"volumes": [0.0] * n, "total": 0.0,
                              "invoiced_volumes": [0.0] * n, "invoiced_total": 0.0},
        }

    ordered = _group_and_sort(plant_rows)
    comp_vols = [round(sum(p["volumes"][i] for p in plant_rows), 2) for i in range(n)]
    comp_inv  = [round(sum(p["invoiced_volumes"][i] for p in plant_rows), 2) for i in range(n)]

    return {
        "type": "monthly",
        "from_month": from_month.isoformat(),
        "to_month": to_month.isoformat(),
        "months": [m.isoformat() for m in months],
        "month_labels": month_labels,
        "regions": ordered,
        "company_total": {
            "volumes": comp_vols,
            "total": round(sum(comp_vols), 2),
            "invoiced_volumes": comp_inv,
            "invoiced_total": round(sum(comp_inv), 2),
        },
    }


# ── Excel export ──────────────────────────────────────────────────────────────

def export_analytics_excel(data: dict, view: str = "production") -> io.BytesIO:
    """Export an analytics result dict (daywise or monthly) to a formatted Excel workbook.

    view: "production" exports produced volumes (data["volumes"]/["total"]);
          "invoiced" exports invoiced quantities (data["invoiced_volumes"]/["invoiced_total"]) —
          matches whichever tab the on-screen toggle was showing when exported.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    is_invoiced = view == "invoiced"
    vol_key, tot_key = ("invoiced_volumes", "invoiced_total") if is_invoiced else ("volumes", "total")

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    is_day = data["type"] == "daywise"
    sheet_label = ("Day-wise" if is_day else "Monthly") + (" Invoiced" if is_invoiced else " Report")
    ws.title = sheet_label[:31]

    # ── Shared styles ──────────────────────────────────────────────────────
    hdr_font  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    reg_font  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    reg_fill  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    tot_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    tot_fill  = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

    dat_font  = Font(name="Calibri", size=10)
    r_align   = Alignment(horizontal="right")
    l_align   = Alignment(horizontal="left")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    col_labels = data.get("date_labels" if is_day else "month_labels", [])
    n_data = len(col_labels)
    period = (
        f"{data['from_date']} to {data['to_date']}"
        if is_day
        else f"{data['from_month']} to {data['to_month']}"
    )

    # Total columns: Sr + Plant Name + [data cols] + Total
    total_cols = 2 + n_data + 1

    # ── Title rows ─────────────────────────────────────────────────────────
    metric_label = "Invoiced Quantity" if is_invoiced else "Production Volume"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1).value = (
        f"Day-wise {metric_label} Report"
        if is_day
        else f"Monthly {metric_label} Summary"
    )
    ws.cell(1, 1).font = Font(name="Calibri", size=14, bold=True)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(2, 1).value = f"Period: {period}"
    ws.cell(2, 1).font = Font(name="Calibri", size=10, italic=True)

    # ── Header row (row 4) ─────────────────────────────────────────────────
    header_vals = ["Sr.", "Plant Name"] + list(col_labels) + [f"Total {metric_label}" if is_invoiced else "Total"]
    for ci, hv in enumerate(header_vals, 1):
        c = ws.cell(4, ci, hv)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    for i in range(n_data):
        ws.column_dimensions[get_column_letter(3 + i)].width = 9
    ws.column_dimensions[get_column_letter(3 + n_data)].width = 12

    # ── Data rows ──────────────────────────────────────────────────────────
    ri = 5
    sr = 1

    for region in data["regions"]:
        for plant in region["plants"]:
            row_data = [sr, plant["plant_name"]] + plant[vol_key] + [plant[tot_key]]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(ri, ci, val)
                c.font = dat_font
                c.alignment = l_align if ci == 2 else r_align
                c.border = thin
            ri += 1
            sr += 1

        # Region subtotal
        sub = region.get("subtotal")
        if sub and region["name"] != "Unassigned":
            sub_data = ["", region["name"]] + sub[vol_key] + [sub[tot_key]]
            for ci, val in enumerate(sub_data, 1):
                c = ws.cell(ri, ci, val)
                c.font = reg_font
                c.fill = reg_fill
                c.alignment = l_align if ci == 2 else r_align
                c.border = thin
            ri += 1

    # Company total
    ct = data["company_total"]
    ct_data = ["", "COMPANY TOTAL"] + ct[vol_key] + [ct[tot_key]]
    for ci, val in enumerate(ct_data, 1):
        c = ws.cell(ri, ci, val)
        c.font = tot_font
        c.fill = tot_fill
        c.alignment = l_align if ci == 2 else r_align
        c.border = thin

    # Freeze header row and plant name column
    ws.freeze_panes = "C5"

    wb.save(output)
    output.seek(0)
    return output
